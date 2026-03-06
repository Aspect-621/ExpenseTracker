import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime


def export_to_excel(transactions_qs, payment_methods_qs):
    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="f8f9fa")
    border      = Border(bottom=Side(style='thin', color='dee2e6'))

    # ── All Transactions ──
    ws = wb.active
    ws.title = "All Transactions"
    headers = ['Date', 'Description', 'Type', 'Amount', 'Currency', 'Exchange Rate', 'Converted Amount', 'From Account', 'To Account', 'Category', 'Notes', 'Hidden']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    for row_idx, t in enumerate(transactions_qs, 2):
        row_data = [
            t.date,
            t.name,
            t.get_transaction_type_display(),
            float(t.amount),
            t.currency,
            float(t.exchange_rate) if t.exchange_rate else '',
            float(t.converted_amount) if t.converted_amount else '',
            t.payment_method.name if t.payment_method else '',
            t.to_payment_method.name if t.to_payment_method else '',
            t.category.name if t.category else '',
            t.notes,
            'Yes' if t.is_hidden else 'No',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical='center')
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if col == 1 and isinstance(val, datetime.date):
                cell.number_format = 'YYYY-MM-DD'
            if col in (4, 7):
                cell.number_format = '#,##0.00'
            if col == 6:
                cell.number_format = '0.0000'

    col_widths = [12, 35, 14, 12, 10, 14, 16, 18, 18, 18, 30, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # ── Balances ──
    from django.db.models import Sum
    ws2 = wb.create_sheet("Balances")
    ws2.append(['Payment Method', 'Currency', 'Starting Balance', 'Current Balance'])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for pm in payment_methods_qs:
        ws2.append([pm.name, pm.currency, float(pm.starting_balance), float(pm.get_balance())])
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18

    # ── By Category ──
    ws3 = wb.create_sheet("By Category")
    ws3.append(['Category', 'Total Spent', 'Currency', 'Count'])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    from apps.expenses.models import Transaction
    by_cat = transactions_qs.filter(transaction_type=Transaction.TYPE_EXPENSE)\
        .values('category__name', 'currency').annotate(total=Sum('amount'), count=Sum('amount') * 0 + 1)\
        .order_by('category__name')
    for row in by_cat:
        ws3.append([row['category__name'], float(row['total']), row['currency'], row['count']])
    for col in ['A','B','C','D']:
        ws3.column_dimensions[col].width = 16

    return wb
