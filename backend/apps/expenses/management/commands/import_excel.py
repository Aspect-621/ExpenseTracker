import datetime
from decimal import Decimal
from django.core.management.base import BaseCommand

VALID_TYPES = {'expense', 'income', 'transfer', 'topup', 'ph_withdrawal', 'tw_withdrawal'}


def parse_row(row):
    """
    Parse a row from the new 10-column import format.
    Columns:
      A(0) Date  B(1) Description  C(2) Type  D(3) From Account
      E(4) To Account  F(5) Category  G(6) Amount  H(7) Currency
      I(8) Discount  J(9) Notes
    Returns a dict or None if the row should be skipped.
    """
    raw_date = row[0]
    name     = row[1]
    txn_type = row[2]
    from_acc = row[3]
    to_acc   = row[4]
    category = row[5]
    amount   = row[6]
    currency = row[7]
    discount = row[8]
    notes    = row[9] if len(row) > 9 else None

    if not name or not txn_type or not from_acc or amount is None:
        return None
    if str(name).strip().lower() in ('description', 'name', ''):
        return None

    txn_type = str(txn_type).strip().lower()
    if txn_type not in VALID_TYPES:
        return None

    if isinstance(raw_date, datetime.datetime):
        txn_date = raw_date.date()
    elif isinstance(raw_date, datetime.date):
        txn_date = raw_date
    elif isinstance(raw_date, str):
        try:
            txn_date = datetime.date.fromisoformat(str(raw_date)[:10])
        except Exception:
            txn_date = datetime.date.today()
    else:
        txn_date = datetime.date.today()

    return {
        'date':     txn_date,
        'name':     str(name).strip(),
        'type':     txn_type,
        'from_acc': str(from_acc).strip(),
        'to_acc':   str(to_acc).strip() if to_acc else None,
        'category': str(category).strip() if category else None,
        'amount':   abs(float(amount)),
        'currency': str(currency).strip() if currency else 'TWD',
        'discount': float(discount) if discount else None,
        'notes':    str(notes).strip() if notes else '',
    }


class Command(BaseCommand):
    help = 'Import transactions from the new 10-column Excel format'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default='/app/import_data/expenses_import_ready.xlsx',
            help='Path to Excel file'
        )
        parser.add_argument('--clear', action='store_true', help='Clear existing transactions before import')

    def handle(self, *args, **options):
        import openpyxl
        from apps.expenses.models import Transaction, PaymentMethod, Category

        filepath = options['file']
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except FileNotFoundError:
            self.stderr.write(f"File not found: {filepath}")
            return

        if options['clear']:
            Transaction.objects.all().delete()
            self.stdout.write("Cleared existing transactions.")

        pm_map  = {pm.name: pm for pm in PaymentMethod.objects.all()}
        cat_map = {c.name: c  for c in Category.objects.all()}
        gen_cat, _ = Category.objects.get_or_create(
            name='General', defaults={'color': '#6c757d', 'icon': 'bag', 'order': 3}
        )

        imported = skipped = errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self.stdout.write(f"\nProcessing: {sheet_name}")

            for raw_row in ws.iter_rows(min_row=4, values_only=True):
                parsed = parse_row(raw_row)
                if not parsed:
                    skipped += 1
                    continue

                from_pm = pm_map.get(parsed['from_acc'])
                if not from_pm:
                    self.stderr.write(f"  Unknown account '{parsed['from_acc']}' — skipping '{parsed['name']}'")
                    errors += 1
                    continue

                to_pm = pm_map.get(parsed['to_acc']) if parsed['to_acc'] else None
                cat   = cat_map.get(parsed['category'], gen_cat) if parsed['category'] else gen_cat

                try:
                    Transaction.objects.create(
                        date=parsed['date'],
                        name=parsed['name'],
                        amount=Decimal(str(parsed['amount'])),
                        currency=parsed['currency'],
                        transaction_type=parsed['type'],
                        payment_method=from_pm,
                        to_payment_method=to_pm,
                        category=cat,
                        discount_amount=Decimal(str(parsed['discount'])) if parsed['discount'] else None,
                        notes=parsed['notes'],
                    )
                    imported += 1
                except Exception as e:
                    self.stderr.write(f"  Error importing '{parsed['name']}': {e}")
                    errors += 1

        self.stdout.write(self.style.SUCCESS(
            f"\n✅ Import complete: {imported} imported, {skipped} skipped, {errors} errors"
        ))
