"""
Clean import script for Taiwan expense data.
Run inside container: python clean_import.py
Reads ONLY: December 2025, January 2026, February 2026
"""
import os, sys, django, datetime
from decimal import Decimal

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'expense_tracker.settings')
sys.path.insert(0, '/app')
django.setup()

from apps.expenses.models import Transaction, PaymentMethod, Category
import openpyxl

FILEPATH = '/app/import_data/Expenses_V2__Taiwan_.xlsx'
SHEETS   = ['December 2025', 'January 2026', 'February 2026']

def classify(name, amount, pm_name):
    n = name.lower().strip()

    if amount > 0:
        # Salary → Income into Taishin Bank
        if 'salary' in n or 'wage' in n:
            return ('income', pm_name, None)

        # Discounts → Income (refund) into same account
        if 'discount' in n or 'refund' in n or 'cashback' in n:
            return ('income', pm_name, None)

        # Withdraw → Transfer from Taishin Bank to Cash
        if 'withdraw' in n:
            return ('transfer', 'Taishin Bank', 'Cash')

        # Converted to Cash → Transfer from current account to Cash
        if 'converted to cash' in n or 'convert to cash' in n:
            return ('transfer', pm_name, 'Cash')

        # Laba Top Up / Top Up on LINE Pay → Transfer from Cash to LINE Pay
        if 'laba' in n or ('top' in n and 'up' in n):
            if pm_name == 'LINE Pay':
                return ('transfer', 'Cash', 'LINE Pay')
            if pm_name == 'iPass':
                return ('transfer', 'Cash', 'iPass')
            return ('income', pm_name, None)

        # TopUp on iPass → Transfer from Cash to iPass
        if n == 'topup':
            return ('transfer', 'Cash', pm_name)

        # Default positive → income
        return ('income', pm_name, None)

    else:
        # Negative = expense
        return ('expense', pm_name, None)


def run():
    wb = openpyxl.load_workbook(FILEPATH, data_only=True)

    pm_map  = {pm.name: pm for pm in PaymentMethod.objects.all()}
    cat_map = {c.name: c  for c in Category.objects.all()}

    general_cat = cat_map.get('General') or Category.objects.get_or_create(
        name='General', defaults={'color': '#6c757d', 'icon': 'bag', 'order': 3})[0]
    income_cat  = cat_map.get('Income') or Category.objects.get_or_create(
        name='Income', defaults={'color': '#198754', 'icon': 'cash-stack', 'order': 4})[0]
    food_cat    = cat_map.get('Food') or Category.objects.get_or_create(
        name='Food', defaults={'color': '#fd7e14', 'icon': 'cart', 'order': 1})[0]

    # CLEAR existing transactions
    count = Transaction.objects.all().count()
    Transaction.objects.all().delete()
    print(f"Cleared {count} existing transactions.")

    imported = errors = 0

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found, skipping.")
            continue

        ws = wb[sheet_name]
        print(f"\nProcessing: {sheet_name}...")

        for row in ws.iter_rows(min_row=4, values_only=True):
            raw_date  = row[0]
            raw_name  = row[1]
            raw_pm    = row[2]
            raw_type  = row[3]
            raw_price = row[4]

            if not raw_name or not raw_pm or raw_price is None:
                continue
            if str(raw_name).strip().lower() in ('name', 'description', ''):
                continue

            # Parse date
            if isinstance(raw_date, datetime.datetime):
                txn_date = raw_date.date()
            elif isinstance(raw_date, datetime.date):
                txn_date = raw_date
            else:
                txn_date = datetime.date.today()

            name     = str(raw_name).strip()
            pm_name  = str(raw_pm).strip()
            cat_name = str(raw_type).strip() if raw_type else 'General'
            amount   = float(raw_price)
            abs_amt  = abs(amount)

            txn_type, from_name, to_name = classify(name, amount, pm_name)

            # Resolve accounts
            from_pm = pm_map.get(from_name)
            to_pm   = pm_map.get(to_name) if to_name else None

            if not from_pm:
                print(f"  ⚠ Unknown account '{from_name}' for '{name}' — skipping")
                errors += 1
                continue

            # Resolve category
            if txn_type == 'income':
                cat = income_cat
            elif txn_type in ('transfer', 'topup'):
                cat = general_cat
            else:
                cat = cat_map.get(cat_name, general_cat)

            try:
                Transaction.objects.create(
                    date=txn_date,
                    name=name,
                    amount=Decimal(str(abs_amt)),
                    currency='TWD',
                    transaction_type=txn_type,
                    payment_method=from_pm,
                    to_payment_method=to_pm,
                    category=cat,
                    notes='',
                    is_hidden=False,
                )
                imported += 1
            except Exception as e:
                print(f"  ✗ Error '{name}': {e}")
                errors += 1

    print(f"\n{'='*50}")
    print(f"✅ Done! {imported} imported, {errors} errors.")
    print(f"{'='*50}")

if __name__ == '__main__':
    run()
